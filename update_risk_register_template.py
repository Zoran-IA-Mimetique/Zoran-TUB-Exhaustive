import argparse
import pandas as pd
from openpyxl import load_workbook

def update_risk_register_template(excel_template_path, csv_file_path, sheet_name, header_row, start_col, out_path, include_headers=False):
    df = pd.read_csv(csv_file_path)
    wb = load_workbook(excel_template_path)
    sh = wb[sheet_name]

    r_offset = 0
    if include_headers:
        for j, col in enumerate(df.columns, start=start_col):
            sh.cell(row=header_row, column=j, value=col)
        r_offset = 1

    for i, row in df.iterrows():
        for j, val in enumerate(row.values, start=start_col):
            sh.cell(row=header_row + r_offset + i, column=j, value=val)

    wb.save(out_path)
    print(f"✅ Excel mis à jour → {out_path}")

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', required=True)
    ap.add_argument('--csv', required=True)
    ap.add_argument('--sheet', required=True)
    ap.add_argument('--header-row', type=int, default=1)
    ap.add_argument('--start-col', type=int, default=1)
    ap.add_argument('--out', required=True)
    ap.add_argument('--include-headers', action='store_true')
    args = ap.parse_args()

    update_risk_register_template(args.excel, args.csv, args.sheet, args.header_row, args.start_col, args.out, args.include_headers)